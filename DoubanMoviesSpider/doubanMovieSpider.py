#-*- coding: UTF-8 -*-

import sys
import time
import urllib3
import certifi
import requests
import numpy as np
from bs4 import BeautifulSoup
from openpyxl import Workbook

#reload(sys)
#sys.setdefaultencoding('utf8')

#Some User Agents
hds=[{'User-Agent':'Mozilla/5.0 (Windows; U; Windows NT 6.1; en-US; rv:1.9.1.6) Gecko/20091201 Firefox/3.5.6'},\
{'User-Agent':'Mozilla/5.0 (Windows NT 6.2) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.12 Safari/535.11'},\
{'User-Agent': 'Mozilla/5.0 (compatible; MSIE 10.0; Windows NT 6.2; Trident/6.0)'}]

def get_hot_movie():
    page_num=0;
    data_type=['片名','主演','导演','豆瓣评分','时长','上映地区']
    movie_list=[]
    try_times=0
    
    #print "电影开始book_spider"+book_tag
    #url='https://www.douban.com/tag/'+urllib.quote(book_tag)+'/book?start='+str(page_num*15)
    #url='https://www.douban.com'
    #url='https://movie.douban.com/tag/#/?sort=U&range=0,10&tags='+book_tag
    url='https://movie.douban.com'
    time.sleep(np.random.rand()*5)
    try:  
        http = urllib3.PoolManager(cert_reqs='CERT_REQUIRED',ca_certs=certifi.where())  # 创建请求
        ret = http.request('GET', url)  # 返回一个 HTTPResponse 对象
        #print(ret.status)  # 返回状态 200/404/...
        #print(ret.data.decode('utf-8')) 
        plain_text = ret.data.decode('utf-8')
        #print plain_text
    except (urllib3.exceptions.HTTPError) as e:
        print("OS error: {0}".format(e))
        
    soup = BeautifulSoup(plain_text)
    list_soup = soup.find('div', {'class': 'screening-bd'})
    #print list_soup
    
    for movie_info in list_soup.findAll('li',{'class':'ui-slide-item'}):
        try:
            title = movie_info.get('data-title')
            actors = movie_info.get('data-actors')
            director = movie_info.get('data-director')
            rate = movie_info.get('data-rate')
            duration = movie_info.get('data-duration')
            region = movie_info.get('data-region')
            
            print ("片名："+title)
            print ("主演："+actors)
            print ("导演："+director)
            print ("豆瓣评分："+rate)
            print ("时长："+duration)
            print ("上映地区："+region)
            print ("===================================================")
            movie_list.append([title,actors,director,rate,duration,region])
        except TypeError:
            print ("这个位置没有海报")
            print ("===================================================")
    
    return movie_list,data_type 

def print_lists_to_excel(movie_lists,data_types):
    #for movie in movie_lists:
    #    print movie[0]
    
    #for atype in data_type:
    #    print atype
    
    wb=Workbook()
    ws=wb.active
    ws.title = '热映片单'
    #ws['A1'] = data_type[0] 
    for i in range(len(data_types)):
        ws.cell(row=1, column=i+1, value=data_types[i])
        for j in range(len(movie_lists)):
            ws.cell(row=j+2, column=i+1, value=movie_lists[j][i])
        
    wb.save('热映片单_'+time.strftime("%Y.%m.%d_%H.%M.%S", time.localtime())+'.xlsx')

if __name__=='__main__':
    movie_lists,data_types=get_hot_movie()
    print_lists_to_excel(movie_lists,data_types)